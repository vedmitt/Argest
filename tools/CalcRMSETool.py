import math
import os
from qgis.core import edit, QgsVectorFileWriter
from qgis.utils import iface

from PyQt5.QtCore import QVariant, QCoreApplication
from PyQt5.QtWidgets import QPushButton, QLineEdit, QToolButton
from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series
)
from qgis._core import QgsProject, QgsVectorLayer, QgsWkbTypes, QgsFeatureRequest, QgsField

from .dataStorage.FeaturesList import FeaturesList
from .dataStorage.FileManager import FileManager
from .dataStorage.FileWriterUtil import FileWriterUtil
from .mathUtil.AzimutMathUtil import AzimutMathUtil
from .mathUtil.InterpolationMathUtil import InterpolationMathUtil


class CalcRMSETool:
    def __init__(self, guiUtil):
        self.guiUtil = guiUtil

    def calcRMSE_test(self, entries):
        lg = FileManager(self.guiUtil)

        # create input layer and extract all features from it to the FeaturesList
        # vlayer = lg.getVlayerFromCanvasByName(self.comboBox_input_2.currentText())
        # content = self.rm_lw_input.selectedItems()
        lyrNames = [i.text() for i in list(entries)]
        vlayer = lg.getVlayerFromCanvasByName(lyrNames[0])
        features = FeaturesList(vlayer.fields().names(), [f.type() for f in vlayer.fields()], vlayer.getFeatures())

        # check existing fields and create new ones if they are not in list
        out_fields = ['T_r', 'T_k']
        flight_nums = [822, 823]
        x_field = 'x'
        t_field = 'T'

        if out_fields not in features.getFieldTypes():
            features.addNewField(out_fields[0], 'Real')
            features.addNewField(out_fields[1], 'Real')

        # lets calculate interpolation
        features.sortByField(x_field)
        tool = InterpolationMathUtil(self.guiUtil)
        # сначала считаем рядовой по Т, затем считаем контрольный по Tr
        features = tool.interpolate_mainFunc(features, flight_nums, t_field, x_field, out_fields)

        N = features.featureCount()
        T1 = features.getVectorOfValues(out_fields[0])  # рядовой
        T2 = features.getVectorOfValues(out_fields[1])  # контрольный
        # guiUtil.setOutputStyle([0, 'N = ' + str(N)])
        # guiUtil.setOutputStyle([0, '\nT1 = ' + str(T1)])
        # guiUtil.setOutputStyle([0, '\nT2 = ' + str(T2)])

        # среднеквадратичное отклонение
        RMSE = tool.calcRMSE(T1, T2)
        self.guiUtil.setOutputStyle([0, '\nс.к.п. = ' + str(RMSE)])

        # среднее отклонение
        MD = tool.calcMeanDeviation(N, T1, T2)
        self.guiUtil.setOutputStyle([0, '\nср. отклонение = ' + str(MD)])

        # стандартное отклонение контрольных и рядовых
        Tr_psd = tool.calcPSD(T1)
        self.guiUtil.setOutputStyle([0, '\nстанд. откл. рядового = ' + str(Tr_psd)])

        Tk_psd = tool.calcPSD(T2)
        self.guiUtil.setOutputStyle([0, '\nстанд. откл. контрольного = ' + str(Tk_psd)])

        # save to the output file
        # save_file_attr = lg.getSaveFileAttr(self.rm_output)
        fpath = r'/Users/ronya/My_Documents/Darhan/controls/plots/823_test.shp'
        mess = lg.createNewShapefile(features, fpath, "UTF-8")
        self.guiUtil.setOutputStyle(mess)

        # # show the plot
        # fpath = r'/Users/ronya/My_Documents/Darhan/controls/plots/823_plot_Tr.shp'
        # fname = 'test_plot_Tr'
        # new_geom = features.getArrayOfValues([x_field, 'T_r'])
        # mess = lg.createNewShapefile(features, fpath, "UTF-8", new_geom)
        #
        # fpath = r'/Users/ronya/My_Documents/Darhan/controls/plots/823_plot_Tk.shp'
        # fname = 'test_plot_Tk'
        # new_geom = features.getArrayOfValues([x_field, 'T_k'])
        # mess = lg.createNewShapefile(features, fpath, "UTF-8", new_geom)

    def main_calcRMSE(self, entries, outDir):
        lg = FileManager()

        # scheme_path = r'/Users/ronya/My_Documents/Darhan/control/scheme.shp'
        # ordinary_path = r'/Users/ronya/My_Documents/Darhan/control/ryadovye_from_sgusch.shp'
        # control_path = r'/Users/ronya/My_Documents/Darhan/control/control_from_sgusch.shp'
        # schemeLyr = QgsVectorLayer(scheme_path, '', 'ogr')
        # out_fields = schemeLyr.out_fields()
        # features = schemeLyr.getFeatures()
        # i = 1
        # flight_num = 800
        # for f in features:
        #     split_scheme = '/Users/ronya/My_Documents/Darhan/extra_controls/scheme_'+str(flight_num+1)+'.shp'
        #     ordinary = '/Users/ronya/My_Documents/Darhan/extra_controls/ordinary_'+str(flight_num+1)+'.shp'
        #     control = '/Users/ronya/My_Documents/Darhan/extra_controls/controls_'+str(flight_num+1)+'.shp'
        #     result = '/Users/ronya/My_Documents/Darhan/extra_controls/control_'+str(flight_num+1)+'.shp'
        #
        #     # lg.createLayerByWriter(out_fields, f, schemeLyr.sourceCrs(), split_scheme)
        #     writer = QgsVectorFileWriter(split_scheme, 'UTF-8', out_fields, QgsWkbTypes.Polygon, schemeLyr.sourceCrs(), 'ESRI Shapefile')
        #     f.setGeometry(f.geometry())
        #     writer.addFeature(f)
        #     del (writer)
        #     # iface.addVectorLayer(split_scheme, '', 'ogr')
        #
        #     # сделаем интерсекцию отдельно рядового и контрольного профилей
        #     lg.intersectionOperation(ordinary_path, split_scheme, ordinary)
        #     lg.intersectionOperation(control_path, split_scheme, control)
        #
        #     # Пронумеруем отдельно контрольный и рядовой профиля
        #     expression = str(flight_num)
        #     lg.createFieldByExpression(ordinary, 'FLIGHT_NUM', QVariant.Int, expression)
        #     # iface.addVectorLayer(ordinary, '', 'ogr')
        #
        #     expression = str(flight_num+1)
        #     lg.createFieldByExpression(control, 'FLIGHT_NUM', QVariant.Int, expression)
        #     # iface.addVectorLayer(control, '', 'ogr')
        #
        #     # Объединим профиля и пронумеруем снова
        #     rlayer = QgsVectorLayer(ordinary, '', 'ogr')
        #     klayer = QgsVectorLayer(control, '', 'ogr')
        #     lg.mergeVectorLayers([rlayer, klayer], result)
        #
        #     expression = str(flight_num+1)
        #     lg.createFieldByExpression(result, 'FLIGHT_N_1', QVariant.Int, expression)
        #     iface.addVectorLayer(result, '', 'ogr')
        #
        #     i += 1
        #     flight_num += 2
        #     # break

        # outDir = self.rm_output.text()
        # # layerName = self.rm_input_1.currentText()
        # content = self.rm_lw_input.selectedItems()
        lyrNames = [i.text() for i in list(entries)]
        if len(lyrNames) > 0:
            sumOfControlsLength = 0
            for layerName in lyrNames:
                sumOfControlsLength = self.calcRMSE_loop(layerName, outDir, sumOfControlsLength)

            # Cохраним итоги в общей таблице
            wb = load_workbook('controls_report.xlsx')
            ws = wb.active
            ws.title = 'Итоговая таблица'
            ws['B2'] = 'Общие результаты'
            ws['B3'] = 'Сумма контрольных профилей'
            ws['C3'] = sumOfControlsLength
            wb.save('controls_report.xlsx')

    def calcRMSE_loop(self, layerName, outDir, sumOfControlsLength):
        lg = FileManager(self.guiUtil)
        az = AzimutMathUtil()
        tool = InterpolationMathUtil(self.guiUtil)
        x_field = 'x'
        y_field = 'y'
        t_field = 'T'

        # whatType = {'txt': self.rb_txt.isChecked(),
        #             'shp': self.rb_shp.isChecked()
        #             }

        userOutdir = outDir
        # создадим папку для результатов по имени папки
        outDir = os.path.join(outDir, layerName)
        if not os.path.exists(outDir):
            os.mkdir(outDir)

        # Проверим является ли геометрия входного слоя MultiPoint
        vlayer = lg.getVlayerFromCanvasByName(layerName)
        # # out_fields = vlayer.out_fields()
        # features = vlayer.getFeatures()
        # t = ''
        # for f in features:
        #     geom = f.geometry()
        #     x, t = lg.getFeatureGeometry(geom)
        #     self.guiUtil.setOutputStyle([0, '\nТип геометрии слоя: ' + t])
        #     break
        # if t == 'MultiPoint':
        #     outFn = os.path.join(outDir, layerName + '.shp')
        #     lg.multipartToSingleparts(vlayer, outFn)
        #     vlayer = QgsVectorLayer(outFn, layerName, 'ogr')
        #     # out_fields = vlayer.out_fields()
        #     # features = vlayer.getFeatures()
        #     # iface.addVectorLayer(outFn, '', 'ogr')

        if vlayer.featureCount() > 0:
            # Выясним в каком типе файлов нужно сохранить результат
            extension = 'shp'
            # if whatType.get('txt'):
            #     extension = 'txt'
            # elif whatType.get('shp'):
            #     extension = 'shp'

            rmse_path = os.path.join(outDir, layerName + '_rmse.txt')
            # affpath = os.path.join(outDir, layerName + '_affine.shp')
            rpath = os.path.join(outDir, layerName + '_ordinary.' + extension)
            kpath = os.path.join(outDir, layerName + '_control.' + extension)
            respath = os.path.join(outDir, layerName + '_result.' + extension)

            linesToSave = []
            linesToSave.append('Data for ' + layerName + '\n')
            FileWriterUtil().writeLinesToFile(rmse_path, linesToSave)

            features = FeaturesList(vlayer.fields().names(), [f.type() for f in vlayer.fields()], vlayer.getFeatures())

            # check existing out_fields and create new ones if they are not in list
            tr_field = 'Yr'  # рядовой, контрольный
            tk_field = 'Yk'

            # Найдем номера рядового и контрольного профилей
            control_num = features.getVectorOfValues("FLIGHT_N_1")
            control_num = list(set(control_num))

            simple_num = features.getVectorOfValues("FLIGHT_NUM")
            simple_num = list(set(simple_num))
            simple_num.remove(control_num[0])
            flight_nums = [int(simple_num[0]), int(control_num[0])]  # [116, 114] рядовой, контрольный

            calculations = []
            linesToSave.append('\nOrdinary #' + str(flight_nums[0]) + ', Control #' + str(flight_nums[1]))
            self.guiUtil.setOutputStyle([0, '\nРядовой №' + str(flight_nums[0]) + ',  контрольный №' + str(flight_nums[1])])
            calculations.append(['Рядовой №', flight_nums[0]])
            calculations.append(['Контрольный №', flight_nums[1]])

            features = tool.interpolate_mainFunc(features, flight_nums, t_field, x_field, [tr_field, tk_field])
            lg.createNewShapefile(features, respath)

            # Выделим отдельно рядовые и контрольные профиля
            expression = '"FLIGHT_NUM" = ' + str(flight_nums[0])
            out_fields, feats = lg.selectByExpression(vlayer, expression)
            lg.createLayerByWriter(out_fields, feats, vlayer.sourceCrs(), rpath, 'point')
            rlayer = QgsVectorLayer(rpath, '', 'ogr')
            r_feats = FeaturesList(rlayer.out_fields().names(), [f.type() for f in rlayer.out_fields()], rlayer.getFeatures())
            # iface.addVectorLayer(rpath, '', 'ogr')

            expression = '"FLIGHT_NUM" = ' + str(flight_nums[1])
            out_fields, feats = lg.selectByExpression(vlayer, expression)
            lg.createLayerByWriter(out_fields, feats, vlayer.sourceCrs(), kpath, 'point')
            klayer = QgsVectorLayer(kpath, '', 'ogr')
            k_feats = FeaturesList(klayer.out_fields().names(), [f.type() for f in klayer.out_fields()], klayer.getFeatures())
            # iface.addVectorLayer(kpath, '', 'ogr')

            # Выравним длину профилей по минимальной длине одного из них
            x0 = features.selectValueByCondition([x_field], ['FLIGHT_NUM', flight_nums[0]])
            y0 = features.selectValueByCondition([y_field], ['FLIGHT_NUM', flight_nums[0]])
            x1 = features.selectValueByCondition([x_field], ['FLIGHT_NUM', flight_nums[1]])
            y1 = features.selectValueByCondition([y_field], ['FLIGHT_NUM', flight_nums[1]])

            minX = max([min(x0), min(x1)])
            minY = max([min(y0), min(y1)])
            maxX = min([max(x0), max(x1)])
            maxY = min([max(y0), max(y1)])

            # self.guiUtil.setOutputStyle([0, '\nminX = ' + str(minX) + ' maxX = ' + str(maxX)])
            # self.guiUtil.setOutputStyle([0, 'minY = ' + str(minY) + ' maxY = ' + str(maxY)])

            # guiUtil.setOutputStyle([0, '\nДлина рядового = ' + str(r_feats.featureCount())])
            # guiUtil.setOutputStyle([0, 'Длина контрольного = ' + str(k_feats.featureCount())])

            r_feats.sortByField(x_field)
            i = 0
            while i < r_feats.featureCount():
                cur_x = int(r_feats.getFeatureValue(i, x_field))
                cur_y = int(r_feats.getFeatureValue(i, y_field))
                if cur_x < minX or cur_x > maxX:
                    # guiUtil.setOutputStyle([0, str(cur_x)])
                    r_feats.removeFeature(i)
                elif cur_y < minY or cur_y > maxY:
                    # guiUtil.setOutputStyle([0, str(cur_y)])
                    r_feats.removeFeature(i)
                i += 1

            k_feats.sortByField(x_field)
            i = 0
            while i < k_feats.featureCount():
                cur_x = int(k_feats.getFeatureValue(i, x_field))
                cur_y = int(k_feats.getFeatureValue(i, y_field))
                if cur_x < minX or cur_x > maxX:
                    # guiUtil.setOutputStyle([0, str(cur_x)])
                    k_feats.removeFeature(i)
                elif cur_y < minY or cur_y > maxY:
                    # guiUtil.setOutputStyle([0, str(cur_y)])
                    k_feats.removeFeature(i)
                i += 1

            self.guiUtil.setOutputStyle([0, '\nДлина рядового = ' + str(r_feats.featureCount())])
            self.guiUtil.setOutputStyle([0, 'Длина контрольного = ' + str(k_feats.featureCount())])

            # посчитаем длину контрольного профиля
            # a = maxX - minX
            # b = maxY - minY
            a = max(x1) - min(x1)
            b = max(y1) - min(y1)
            c = math.sqrt(a ** 2 + b ** 2)
            linesToSave.append('\nLength (m) = ' + str(c))
            self.guiUtil.setOutputStyle([0, '\nДлина контрольного профиля (м) = ' + str(c)])
            calculations.append(['Длина контрольного профиля (м)', c])
            sumOfControlsLength += c
            #
            # # check existing out_fields and create new ones if they are not in list
            # if tr_field not in r_feats.getFields() and tk_field not in k_feats.getFields():
            #     r_feats.addNewField(tr_field, 'Real')
            #     r_feats.addNewField(tk_field, 'Real')
            #
            #     k_feats.addNewField(tr_field, 'Real')
            #     k_feats.addNewField(tk_field, 'Real')
            #
            # # сначала интерполируем рядовой по Т, затем интерполируем контрольный по Tr
            # X1 = r_feats.getVectorOfValues(x_field)
            # X2 = k_feats.getVectorOfValues(x_field)
            #
            # Y = k_feats.getVectorOfValues(t_field)
            # Yr1 = tool.calcInterpolation(X1, X2, Y)
            # Yk2 = tool.calcInterpolation(X2, X1, Yr1)
            # Yr2 = tool.calcInterpolation(X2, X1, Yk2)  # !!
            # Yk1 = tool.calcInterpolation(X1, X2, Yr2)
            #
            # r_feats.setValuesByField(tr_field, Yr1)  # 116 Tr
            # r_feats.setValuesByField(tk_field, Yk1)  # 116 Tk
            # k_feats.setValuesByField(tr_field, Yr2)  # 114 Tr
            # k_feats.setValuesByField(tk_field, Yk2)  # 114 Tk
            # # --------------------------------------------------------------------------
            types = ['integer', 'double', 'double', 'double']
            out_fields = ['FLIGHT_NUM', x_field, tr_field, tk_field]
            records = features.getArrayOfValues(out_fields)
            # records = r_feats.getArrayOfValues(out_fields)
            one_feat_list = FeaturesList(out_fields, types, records)
            # one_feat_list.addNewFeatsToEndOfList(k_feats.getArrayOfValues(out_fields))
            one_feat_list.sortByField(x_field)

            N = one_feat_list.featureCount()
            y_actual = one_feat_list.getVectorOfValues(tr_field)  # рядовой
            y_predicted = one_feat_list.getVectorOfValues(tk_field)  # контрольный
            # y_actual = one_feat_list.selectValueByCondition([tr_field], ['FLIGHT_NUM', flight_nums[0]])
            # y_predicted = one_feat_list.selectValueByCondition([tk_field], ['FLIGHT_NUM', flight_nums[1]])
            # N = len(y_actual)

            self.guiUtil.setOutputStyle([0, '\ny_actual = ' + str(len(y_actual))])
            self.guiUtil.setOutputStyle([0, '\ny_actual = ' + str(y_actual)])
            self.guiUtil.setOutputStyle([0, '\ny_predicted = ' + str(len(y_predicted))])
            self.guiUtil.setOutputStyle([0, '\ny_predicted = ' + str(y_predicted)])

            # среднеквадратичное отклонение
            RMSE = tool.calcRMSE(y_actual, y_predicted)
            linesToSave.append('\nRoot Mean Square Error (RMSE) = ' + str(RMSE))
            self.guiUtil.setOutputStyle([0, '\nс.к.п. = ' + str(RMSE)])
            calculations.append(['с.к.о.', RMSE])

            # среднее отклонение
            MD = tool.calcMeanDeviation(N, y_actual, y_predicted)
            linesToSave.append('\nMean Deviation = ' + str(MD))
            self.guiUtil.setOutputStyle([0, 'ср. отклонение = ' + str(MD)])
            calculations.append(['ср. отклонение', MD])

            # стандартное отклонение контрольных и рядовых
            Tr_psd = tool.calcPSD(y_actual)
            linesToSave.append('\nPopulation Standard Deviation (ordinary) = ' + str(Tr_psd))
            self.guiUtil.setOutputStyle([0, 'станд. откл. рядового = ' + str(Tr_psd)])
            calculations.append(['станд. откл. рядового', Tr_psd])

            Tk_psd = tool.calcPSD(y_predicted)
            linesToSave.append('\nPopulation Standard Deviation (control) = ' + str(Tk_psd))
            self.guiUtil.setOutputStyle([0, 'станд. откл. контрольного = ' + str(Tk_psd)])
            calculations.append(['станд. откл. контрольного', Tk_psd])

            # Сохраним данные в текстовый файл
            FileWriterUtil().writeLinesToFile(rmse_path, linesToSave)

            # Cохраним полученные данные в excell. Каждый профиль на отдельном листе вместе со всей информацией.
            one_feat_list.sortByField(x_field)
            yr_values = one_feat_list.selectValueByCondition([x_field, tr_field], ['FLIGHT_NUM', flight_nums[0]])
            yk_values = one_feat_list.selectValueByCondition([tk_field], ['FLIGHT_NUM', flight_nums[1]])

            os.chdir(userOutdir)
            dest_filename = os.path.join(userOutdir, 'controls_report.xlsx')

            if os.path.exists(dest_filename):
                wb = load_workbook('controls_report.xlsx')
            else:
                wb = Workbook()
            wb.template = False
            ws = wb.create_sheet(title=layerName)
            # ws = wb.active
            # ws.title = layerName
            _ = ws.cell(column=1, row=1, value='X')
            _ = ws.cell(column=2, row=1, value='Рядовые измерения')
            for row in range(2, len(yr_values) + 2):
                for col in range(1, 3):
                    _ = ws.cell(column=col, row=row, value=yr_values[row - 2][col - 1])

            _ = ws.cell(column=3, row=1, value='Контрольные измерения')
            for row in range(2, len(yk_values) + 2):
                _ = ws.cell(column=3, row=row, value=yk_values[row - 2])

            # запишем данные ско и номера профилей
            _ = ws.cell(column=7, row=2, value='Характеристики')
            for row in range(3, len(calculations) + 3):
                for col in range(7, 9):
                    _ = ws.cell(column=col, row=row, value=calculations[row - 3][col - 7])

            # Построим график
            chart = ScatterChart()
            chart.title = "Контрольный профиль №" + str(flight_nums[1]) + " (СКО =" + str(format(RMSE)) + ")"
            # chart.style = 3
            chart.scatterStyle = 'marker'
            chart.x_axis.title = 'X'
            chart.y_axis.title = 'Y'

            xvalues = Reference(ws, min_col=1, min_row=2, max_row=len(yk_values) + 1)
            for i in range(2, 4):
                values = Reference(ws, min_col=i, min_row=1, max_row=len(yk_values) + 1)
                series = Series(values, xvalues, title_from_data=True)
                series.marker = openpyxl.chart.marker.Marker('dot')
                series.graphicalProperties.line.noFill = True
                chart.series.append(series)

            ws.add_chart(chart, "D10")

            wb.save(filename=dest_filename)
            # wb.save('controls_report.xlsx')
            return sumOfControlsLength