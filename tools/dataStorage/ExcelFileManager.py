import os

from openpyxl import Workbook, load_workbook
from openpyxl.chart import (
    ScatterChart,
    Reference,
    Series,
)

class ExcelFileManager:
    def __init__(self, fpath, isOverride):
        self.fpath = fpath
        self.wb = self.createWB(fpath, isOverride)
        self.ws = self.wb.active

    def createWB(self, fpath, isOverride=True):
        if os.path.exists(fpath) and not isOverride:
            wb = load_workbook(fpath)
        else:
            wb = Workbook()
        return wb

    def writeData(self, ind, values):
        minRow = ind[0]
        minCol = ind[1]

        if isinstance(values, list):
            maxRow = ind[2]
            for row in range(minRow, maxRow):
                if isinstance(values[0], list):
                    maxCol = ind[3]
                    for col in range(minCol, maxCol):
                        _ = self.ws.cell(column=col, row=row, value=values[row-minRow][col-minCol])
                else:  # в случае если столбец данных один
                    _ = self.ws.cell(column=minCol, row=row, value=values[row-minRow])
        else:  # в случае если данные только для одной ячейки
            _ = self.ws.cell(column=minCol, row=minRow, value=values)

    def createScatterChart(self, params):
        chart = ScatterChart()
        chart.title = params.get('chart_title')
        chart.scatterStyle = params.get('scatterStyle')
        chart.x_axis.title = params.get('x_axis.title')
        chart.y_axis.title = params.get('y_axis.title')

        x_ind = params.get('x_val.indexes')  # minRow, minCol, maxRow, maxCol
        xvalues = Reference(self.ws, min_col=x_ind[1], min_row=x_ind[0], max_row=x_ind[2])

        y_ind = params.get('y_val.indexes')  # minRow, minCol, maxRow, maxCol
        for i in range(y_ind[1], y_ind[3]):
            values = Reference(self.ws, min_col=i, min_row=y_ind[0], max_row=y_ind[2])
            series = Series(values, xvalues, title_from_data=True)
            series.marker = openpyxl.chart.marker.Marker(params.get('marker_type'))
            series.graphicalProperties.line.noFill = True
            chart.series.append(series)

        self.ws.add_chart(chart, params.get('chart_place'))
        self.wb.save(filename=self.fpath)

    def createSheetWithData(self, sheetName, data):
        self.wb.template = False
        # self.ws = self.wb.create_sheet(title=sheetName)
        # self.ws = self.wb.active
        self.ws.title = sheetName

        for r in data:
            indexes = r[0]
            values = r[1]
            self.writeData(indexes, values)

        # self.writeData(ws, [1, 1], 'X')
        # self.writeData(ws, [1, 2], 'Y')
        #
        # x_values = [0, 1, 2, 3, 4, 5, 6, 7, 8, 9]
        # xindexes = [2, 1, 12]
        # self.writeData(ws, xindexes, x_values)
        #
        # y_values = [0.1, 1.2, 2.3, 3.4, 4.5, 5.6, 6.7, 7.8, 8.9, 9.1]
        # yindexes = [2, 2, 12, 3]
        # self.writeData(ws, yindexes, y_values)

        self.wb.save(filename=self.fpath)
        # wb.save('controls_report.xlsx')

if __name__ == '__main__':
    filepath = '/Users/ronya/My_Documents/output/test.xlsx'
    wb = ExcelFileManager(filepath, True)
    # wb.createSheetWithDataAndPlot('new_test', None, filepath)